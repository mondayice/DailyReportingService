#!/usr/bin/env python
# -*- coding: UTF-8 -*-
'''
@Project ：DailyReportingService
@File    ：main.py
@Author  ：Mondayice
@Date    ：2023年11月7日
'''


from __future__ import print_function

import ctypes
import os
import sys
from datetime import datetime, timedelta

import openpyxl
from openpyxl.styles import Font, Alignment

name = ""

def work():
    # 获取昨天的日期
    yesterday = datetime.now() - timedelta(days=1)
    yesterday_str = yesterday.strftime('%Y-%m-%d')
    # 获取今天的日期
    today = datetime.now()
    today_str = today.strftime('%Y-%m-%d')

    # 询问用户是昨天的日清还是今天的日清
    while True:
        name = input("请输入姓名：")
        date_type = input("请输入日期类型（昨天/今天）：")
        if date_type == '昨天':
            date_str = yesterday_str
            break
        elif date_type == '今天':
            date_str = today_str
            break
        else:
            print("输入错误，请重新输入")

    # 询问用户输入自定义条数
    while True:
        try:
            merge_count = int(input("请输入自定义条数："))
            if merge_count <= 0:
                print("输入错误，请重新输入")
            else:
                break
        except ValueError:
            print("输入错误，请重新输入")

    # 创建一个新的workbook
    wb = openpyxl.Workbook()
    sheet = wb.active
    # 设置 A 列到 G 列的列宽
    sheet.column_dimensions['A'].width = 13.25
    sheet.column_dimensions['B'].width = 41.88
    sheet.column_dimensions['C'].width = 13.25
    sheet.column_dimensions['D'].width = 34.25
    sheet.column_dimensions['E'].width = 29.88
    sheet.column_dimensions['F'].width = 14.88
    sheet.column_dimensions['G'].width = 13.13

    # 对A1到G1的单元格进行合并
    sheet.merge_cells('A1:G1')
    # 设置合并后单元格的值和样式
    merged_cell = sheet['A1']
    merged_cell.value = '日清'
    merged_cell.font = Font(name='微软雅黑', size=18, bold=True)
    merged_cell.alignment = Alignment(horizontal='center', vertical='center')

    # 对A2到G2的单元格进行合并
    sheet.merge_cells('A2:G2')
    # 设置合并后单元格的值和样式
    merged_cell = sheet['A2']
    merged_cell.value = f"姓名：{name} 日期：{date_str}"
    merged_cell.font = Font(name='仿宋', size=11, bold=True)
    merged_cell.alignment = Alignment(horizontal='right', vertical='bottom')

    # 设置从 B3 到 G3 的内容
    headers = ["项目", "目标", "实际", "下一步计划", "备注", "明日工作"]
    for col, header in zip(range(2, 8), headers):
        cell = sheet.cell(row=3, column=col, value=header)

    # 设置从 B3 到 G3 的格式
    for row in range(3, 4):  # 3到4是为了包含第三行
        for col in range(2, 8):
            cell = sheet.cell(row=row, column=col)
            cell.font = openpyxl.styles.Font(name='微软雅黑', size=16, bold=True)
            cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

    # 获取用户输入的合并数量
    #  merge_count = int(input("请输入自定义条数："))
    # 从 A3 开始向下合并单元格
    merged_cell = sheet['A3']
    start_cell = 'A3'
    end_cell = 'A' + str(3 + merge_count)  # 根据合并数量确定结束单元格
    merged_cell.font = Font(name='微软雅黑', size=16, bold=True)
    merged_cell.alignment = Alignment(horizontal='center', vertical='center')
    sheet.merge_cells(start_cell + ':' + end_cell)
    merged_cell.value = "工作内容"

    # 设置从 A 列到 G 列的表格线为黑色
    for col in range(1, 8):
        for row in range(1, merge_count + 4):
            cell = sheet.cell(row=row, column=col)
            cell.border = openpyxl.styles.Border(left=openpyxl.styles.Side(style='thin', color='000000'),
                                                 right=openpyxl.styles.Side(style='thin', color='000000'),
                                                 top=openpyxl.styles.Side(style='thin', color='000000'),
                                                 bottom=openpyxl.styles.Side(style='thin', color='000000'))
    # 合并单元格并设置格式
    merge_range = f'G4:G{4 + merge_count - 1}'  # 计算合并的范围
    sheet.merge_cells(merge_range)
    merged_cell = sheet.cell(row=4, column=7, value=f' ')
    merged_cell.font = openpyxl.styles.Font(name='宋体', size=11)
    merged_cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center', wrap_text=True)

    project_list = {}
    mete = ["项目", "目标", "实际完成情况", "下一步计划", "备注"]
    for i in range(merge_count):
        array = []
        for j in range(5):
            element = input("请输入第{}组的{}：".format(i + 1, mete[j]))
            array.append(element)
        project_list[str(i)] = array
    merged_cell = sheet['G4']
    merged_cell.value = input("明日工作：")
    # 填充数据并设置格式
    for row, i in zip(range(4, merge_count + 4), range(merge_count)):  # 从第四行开始填充数据
        for col, header in zip(range(2, 7), project_list[str(i)]):  # B 列到 F 列
            # cell = sheet.cell(row=row, column=col, value=f'Data{row - 3}-{col - 1}')
            cell = sheet.cell(row=row, column=col, value=header)
            cell.font = openpyxl.styles.Font(name='宋体', size=10.5)
            cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center', wrap_text=True)
    # # 设置全部表格为自动换行
    # for row in sheet.iter_rows():
    #     for cell in row:
    #         cell.alignment = openpyxl.styles.Alignment(wrap_text=True)
    # 保存workbook到文件
    # 保存Excel文件到桌面
    desktop_path = os.path.join(os.path.expanduser("~"), "Desktop")
    file_name = f"日清_{date_str}.xlsx"
    file_path = os.path.join(desktop_path, file_name)
    wb.save(file_path)
    print(f"Excel文件已保存到：{file_path}")


def is_admin():
    try:
        return ctypes.windll.shell32.IsUserAnAdmin()
    except:
        return False


if __name__ == '__main__':
    if is_admin():
        work()
    else:
        if sys.version_info[0] == 3:
            ctypes.windll.shell32.ShellExecuteW(None, "runas", sys.executable, __file__, None, 1)
        else:  # in python2.x
            ctypes.windll.shell32.ShellExecuteW(None, u"runas", unicode(sys.executable), unicode(__file__), None, 1)
